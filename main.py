import requests
from bs4 import BeautifulSoup
import json
import openpyxl


url_1 = 'https://profil-opt.ru/'
wb = openpyxl.Workbook()
ws = wb.active
headers = {
    'Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
}
result = []
link_product_list = []
list_link = []


req = requests.get(url=url_1, headers=headers, verify=False)

soup = BeautifulSoup(req.text,'lxml')

links = soup.find_all(class_='mainmenu2')
for i in links:
    link = i.find('a').get('href')
    list_link.append('https://profil-opt.ru/' + link)
print(list_link)
for product in list_link:
    req_2 = requests.get(product,headers=headers,verify=False)
    soup = BeautifulSoup(req_2.text,'lxml')
    link_product = soup.find_all(class_='shrtstrycat')
    for i in link_product:
        link_product_list.append(i.find('a').get('href'))

for material in link_product_list:
    req_3 = requests.get(material,headers=headers,verify=False)
    soup = BeautifulSoup(req_3.text,'lxml')

    name_material = soup.find_all(itemprop='name')[2]
    for i in name_material:
        name_material_list = i.text

    description_material = soup.find_all(itemprop='description')[1:3]
    for i in description_material:
        description_material_list = i.text.strip()


    id__material = soup.find_all(itemprop='sku')
    for i in id__material:
        id__material_list = i.text

    result.append({
            'name': name_material_list,
            'description':description_material_list,
            'id': id__material_list,
            'url': material
        })
    ws['A1'] = 'Название'
    ws['B1'] = 'Описание'
    ws['C1'] = 'ID'
    ws['D1'] = 'URL'

    for idx, data in enumerate(result, start=2):
        ws[f'A{idx}'] = data['name']
        ws[f'B{idx}'] = data['description']
        ws[f'C{idx}'] = data['id']
        ws[f'D{idx}'] = data['url']

# Сохраняем книгу в файл
wb.save('profil-opt.ru/данные.xlsx')
